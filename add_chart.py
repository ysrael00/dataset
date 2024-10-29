from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

# ler pasta de trabalho e planilha
wb = load_workbook("dataset/pivot_table.xlsx")
sheet = wb["relatorio"]

# referencia das linhas e colunas
min_column = sheet.min_column
max_column = sheet.max_column
print(min_column)
print(max_column)
min_row = sheet.min_row
max_row = sheet.max_row

# adicionando dados e categorias no grafico como valor maximo das linhas e colunas e adicionando categorias e titulos
barchart = BarChart()

data = Reference(
    sheet,
    min_col=min_column + 1,
    max_col=max_column,
    min_row=min_row,
    max_row=max_row
)

categories = Reference(
    sheet,
    min_col=min_column,
    max_col=min_column,
    min_row=min_row + 1,
    max_row=max_row
)

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

# criando o grafico
sheet.add_chart(barchart, "B10")
barchart.title = "Vendas por Fabricantes"
barchart.style = 2

# salvando o workbook
wb.save("dataset/barchart.xlsx")
